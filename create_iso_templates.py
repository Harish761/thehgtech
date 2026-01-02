#!/usr/bin/env python3
"""
Generate ISO 27001:2022 Excel Templates for TheHGTech
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Color scheme
HEADER_FILL = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000", size=11)
SECTION_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
THEME_ORG = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
THEME_PEOPLE = PatternFill(start_color="30D158", end_color="30D158", fill_type="solid")
THEME_PHYSICAL = PatternFill(start_color="FF9500", end_color="FF9500", fill_type="solid")
THEME_TECH = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_gap_analysis_template():
    """Create ISO 27001:2022 Gap Analysis Template"""
    wb = Workbook()
    
    # ========== Sheet 1: Control Assessment ==========
    ws1 = wb.active
    ws1.title = "Control Assessment"
    
    # Headers
    headers = ["Control ID", "Control Name", "Theme", "Implementation Status", 
               "Evidence Quality", "Gap Severity", "Remediation Effort", 
               "Current Evidence", "Gap Description", "Remediation Action", "Owner", "Target Date"]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
    
    set_column_widths(ws1, [12, 40, 15, 20, 18, 15, 18, 30, 30, 35, 15, 12])
    
    # Annex A Controls (93 controls)
    controls = [
        # Organizational Controls (37)
        ("A.5.1", "Policies for information security", "Organizational"),
        ("A.5.2", "Information security roles and responsibilities", "Organizational"),
        ("A.5.3", "Segregation of duties", "Organizational"),
        ("A.5.4", "Management responsibilities", "Organizational"),
        ("A.5.5", "Contact with authorities", "Organizational"),
        ("A.5.6", "Contact with special interest groups", "Organizational"),
        ("A.5.7", "Threat intelligence", "Organizational"),
        ("A.5.8", "Information security in project management", "Organizational"),
        ("A.5.9", "Inventory of information and other associated assets", "Organizational"),
        ("A.5.10", "Acceptable use of information and other associated assets", "Organizational"),
        ("A.5.11", "Return of assets", "Organizational"),
        ("A.5.12", "Classification of information", "Organizational"),
        ("A.5.13", "Labelling of information", "Organizational"),
        ("A.5.14", "Information transfer", "Organizational"),
        ("A.5.15", "Access control", "Organizational"),
        ("A.5.16", "Identity management", "Organizational"),
        ("A.5.17", "Authentication information", "Organizational"),
        ("A.5.18", "Access rights", "Organizational"),
        ("A.5.19", "Information security in supplier relationships", "Organizational"),
        ("A.5.20", "Addressing information security within supplier agreements", "Organizational"),
        ("A.5.21", "Managing information security in the ICT supply chain", "Organizational"),
        ("A.5.22", "Monitoring, review and change management of supplier services", "Organizational"),
        ("A.5.23", "Information security for use of cloud services", "Organizational"),
        ("A.5.24", "Information security incident management planning and preparation", "Organizational"),
        ("A.5.25", "Assessment and decision on information security events", "Organizational"),
        ("A.5.26", "Response to information security incidents", "Organizational"),
        ("A.5.27", "Learning from information security incidents", "Organizational"),
        ("A.5.28", "Collection of evidence", "Organizational"),
        ("A.5.29", "Information security during disruption", "Organizational"),
        ("A.5.30", "ICT readiness for business continuity", "Organizational"),
        ("A.5.31", "Legal, statutory, regulatory and contractual requirements", "Organizational"),
        ("A.5.32", "Intellectual property rights", "Organizational"),
        ("A.5.33", "Protection of records", "Organizational"),
        ("A.5.34", "Privacy and protection of PII", "Organizational"),
        ("A.5.35", "Independent review of information security", "Organizational"),
        ("A.5.36", "Compliance with policies, rules and standards for information security", "Organizational"),
        ("A.5.37", "Documented operating procedures", "Organizational"),
        # People Controls (8)
        ("A.6.1", "Screening", "People"),
        ("A.6.2", "Terms and conditions of employment", "People"),
        ("A.6.3", "Information security awareness, education and training", "People"),
        ("A.6.4", "Disciplinary process", "People"),
        ("A.6.5", "Responsibilities after termination or change of employment", "People"),
        ("A.6.6", "Confidentiality or non-disclosure agreements", "People"),
        ("A.6.7", "Remote working", "People"),
        ("A.6.8", "Information security event reporting", "People"),
        # Physical Controls (14)
        ("A.7.1", "Physical security perimeters", "Physical"),
        ("A.7.2", "Physical entry", "Physical"),
        ("A.7.3", "Securing offices, rooms and facilities", "Physical"),
        ("A.7.4", "Physical security monitoring", "Physical"),
        ("A.7.5", "Protecting against physical and environmental threats", "Physical"),
        ("A.7.6", "Working in secure areas", "Physical"),
        ("A.7.7", "Clear desk and clear screen", "Physical"),
        ("A.7.8", "Equipment siting and protection", "Physical"),
        ("A.7.9", "Security of assets off-premises", "Physical"),
        ("A.7.10", "Storage media", "Physical"),
        ("A.7.11", "Supporting utilities", "Physical"),
        ("A.7.12", "Cabling security", "Physical"),
        ("A.7.13", "Equipment maintenance", "Physical"),
        ("A.7.14", "Secure disposal or re-use of equipment", "Physical"),
        # Technological Controls (34)
        ("A.8.1", "User endpoint devices", "Technological"),
        ("A.8.2", "Privileged access rights", "Technological"),
        ("A.8.3", "Information access restriction", "Technological"),
        ("A.8.4", "Access to source code", "Technological"),
        ("A.8.5", "Secure authentication", "Technological"),
        ("A.8.6", "Capacity management", "Technological"),
        ("A.8.7", "Protection against malware", "Technological"),
        ("A.8.8", "Management of technical vulnerabilities", "Technological"),
        ("A.8.9", "Configuration management", "Technological"),
        ("A.8.10", "Information deletion", "Technological"),
        ("A.8.11", "Data masking", "Technological"),
        ("A.8.12", "Data leakage prevention", "Technological"),
        ("A.8.13", "Information backup", "Technological"),
        ("A.8.14", "Redundancy of information processing facilities", "Technological"),
        ("A.8.15", "Logging", "Technological"),
        ("A.8.16", "Monitoring activities", "Technological"),
        ("A.8.17", "Clock synchronization", "Technological"),
        ("A.8.18", "Use of privileged utility programs", "Technological"),
        ("A.8.19", "Installation of software on operational systems", "Technological"),
        ("A.8.20", "Networks security", "Technological"),
        ("A.8.21", "Security of network services", "Technological"),
        ("A.8.22", "Segregation of networks", "Technological"),
        ("A.8.23", "Web filtering", "Technological"),
        ("A.8.24", "Use of cryptography", "Technological"),
        ("A.8.25", "Secure development life cycle", "Technological"),
        ("A.8.26", "Application security requirements", "Technological"),
        ("A.8.27", "Secure system architecture and engineering principles", "Technological"),
        ("A.8.28", "Secure coding", "Technological"),
        ("A.8.29", "Security testing in development and acceptance", "Technological"),
        ("A.8.30", "Outsourced development", "Technological"),
        ("A.8.31", "Separation of development, test and production environments", "Technological"),
        ("A.8.32", "Change management", "Technological"),
        ("A.8.33", "Test information", "Technological"),
        ("A.8.34", "Protection of information systems during audit testing", "Technological"),
    ]
    
    theme_fills = {
        "Organizational": THEME_ORG,
        "People": THEME_PEOPLE,
        "Physical": THEME_PHYSICAL,
        "Technological": THEME_TECH
    }
    
    for row, (ctrl_id, ctrl_name, theme) in enumerate(controls, 2):
        ws1.cell(row=row, column=1, value=ctrl_id).border = THIN_BORDER
        ws1.cell(row=row, column=2, value=ctrl_name).border = THIN_BORDER
        theme_cell = ws1.cell(row=row, column=3, value=theme)
        theme_cell.border = THIN_BORDER
        theme_cell.fill = theme_fills.get(theme, HEADER_FILL)
        theme_cell.font = Font(color="FFFFFF", bold=True)
        for col in range(4, 13):
            ws1.cell(row=row, column=col).border = THIN_BORDER
    
    # Add data validation dropdowns
    from openpyxl.worksheet.datavalidation import DataValidation
    
    status_dv = DataValidation(type="list", formula1='"Not Implemented,Partially Implemented,Fully Implemented,Not Applicable"')
    evidence_dv = DataValidation(type="list", formula1='"None,Informal,Documented,Proven Effective"')
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    effort_dv = DataValidation(type="list", formula1='"Quick Win,Moderate,Significant Project"')
    
    ws1.add_data_validation(status_dv)
    ws1.add_data_validation(evidence_dv)
    ws1.add_data_validation(severity_dv)
    ws1.add_data_validation(effort_dv)
    
    status_dv.add(f"D2:D{len(controls)+1}")
    evidence_dv.add(f"E2:E{len(controls)+1}")
    severity_dv.add(f"F2:F{len(controls)+1}")
    effort_dv.add(f"G2:G{len(controls)+1}")
    
    # ========== Sheet 2: Gap Scoring Matrix ==========
    ws2 = wb.create_sheet("Gap Scoring Matrix")
    
    headers2 = ["Control ID", "Risk Impact (1-5)", "Likelihood (1-5)", 
                "Audit Probability (1-3)", "Implementation Effort (1-5)", 
                "Priority Score", "Priority Level"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws2, [12, 18, 18, 20, 22, 15, 15])
    
    for row, (ctrl_id, _, _) in enumerate(controls, 2):
        ws2.cell(row=row, column=1, value=ctrl_id).border = THIN_BORDER
        for col in range(2, 6):
            ws2.cell(row=row, column=col).border = THIN_BORDER
        # Formula for Priority Score
        score_cell = ws2.cell(row=row, column=6, value=f"=(B{row}*C{row})+D{row}-E{row}")
        score_cell.border = THIN_BORDER
        # Formula for Priority Level
        level_cell = ws2.cell(row=row, column=7, value=f'=IF(F{row}>=10,"Critical",IF(F{row}>=6,"High",IF(F{row}>=3,"Medium","Low")))')
        level_cell.border = THIN_BORDER
    
    # ========== Sheet 3: Remediation Tracker ==========
    ws3 = wb.create_sheet("Remediation Tracker")
    
    headers3 = ["Control ID", "Gap Description", "Remediation Action", "Owner", 
                "Start Date", "Target Date", "Actual Completion", "Status", 
                "Evidence Link", "Notes"]
    
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws3, [12, 35, 35, 15, 12, 12, 15, 12, 25, 30])
    
    status_tracker = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,Completed,Verified"')
    ws3.add_data_validation(status_tracker)
    status_tracker.add("H2:H200")
    
    # ========== Sheet 4: Evidence Checklist ==========
    ws4 = wb.create_sheet("Evidence Checklist")
    
    headers4 = ["Control ID", "Control Name", "Required Evidence", "Evidence Location", "Collected", "Reviewer Notes"]
    
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws4, [12, 40, 45, 30, 12, 35])
    
    # Sample evidence requirements
    evidence_examples = {
        "A.5.1": "Information Security Policy (approved, dated), evidence of communication to staff",
        "A.5.7": "Threat intelligence sources list, analysis reports, briefing records",
        "A.6.3": "Training program, completion records, assessment results, phishing simulation reports",
        "A.8.9": "Configuration baselines, hardening guides, drift detection reports",
        "A.8.13": "Backup policy, backup logs, restore test records",
        "A.8.16": "SIEM configuration, alert rules, log review records, incident tickets",
    }
    
    row = 2
    for ctrl_id, ctrl_name, _ in controls:
        ws4.cell(row=row, column=1, value=ctrl_id).border = THIN_BORDER
        ws4.cell(row=row, column=2, value=ctrl_name).border = THIN_BORDER
        ws4.cell(row=row, column=3, value=evidence_examples.get(ctrl_id, "Document evidence of implementation")).border = THIN_BORDER
        for col in range(4, 7):
            ws4.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    collected_dv = DataValidation(type="list", formula1='"Yes,No,Partial"')
    ws4.add_data_validation(collected_dv)
    collected_dv.add(f"E2:E{len(controls)+1}")
    
    # ========== Sheet 5: 2013 to 2022 Mapping ==========
    ws5 = wb.create_sheet("2013-2022 Mapping")
    
    headers5 = ["ISO 27001:2022", "2022 Control Name", "ISO 27001:2013", "Change Type"]
    
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws5, [18, 45, 25, 15])
    
    # New controls in 2022
    new_controls = [
        ("A.5.7", "Threat intelligence", "-", "NEW"),
        ("A.5.23", "Information security for use of cloud services", "-", "NEW"),
        ("A.5.30", "ICT readiness for business continuity", "-", "NEW"),
        ("A.7.4", "Physical security monitoring", "-", "NEW"),
        ("A.8.9", "Configuration management", "-", "NEW"),
        ("A.8.10", "Information deletion", "-", "NEW"),
        ("A.8.11", "Data masking", "-", "NEW"),
        ("A.8.12", "Data leakage prevention", "-", "NEW"),
        ("A.8.16", "Monitoring activities", "-", "NEW"),
        ("A.8.23", "Web filtering", "-", "NEW"),
        ("A.8.28", "Secure coding", "-", "NEW"),
    ]
    
    for row, (ctrl_2022, name, ctrl_2013, change) in enumerate(new_controls, 2):
        ws5.cell(row=row, column=1, value=ctrl_2022).border = THIN_BORDER
        ws5.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws5.cell(row=row, column=3, value=ctrl_2013).border = THIN_BORDER
        change_cell = ws5.cell(row=row, column=4, value=change)
        change_cell.border = THIN_BORDER
        change_cell.fill = PatternFill(start_color="30D158", end_color="30D158", fill_type="solid")
        change_cell.font = Font(color="FFFFFF", bold=True)
    
    # ========== Sheet 6: Executive Summary ==========
    ws6 = wb.create_sheet("Executive Summary")
    
    ws6.cell(row=1, column=1, value="ISO 27001:2022 Gap Analysis Executive Summary").font = Font(bold=True, size=16)
    ws6.cell(row=3, column=1, value="Organization:").font = Font(bold=True)
    ws6.cell(row=4, column=1, value="Assessment Date:").font = Font(bold=True)
    ws6.cell(row=5, column=1, value="Assessor:").font = Font(bold=True)
    ws6.cell(row=6, column=1, value="Scope:").font = Font(bold=True)
    
    ws6.cell(row=8, column=1, value="Summary by Theme").font = Font(bold=True, size=14)
    
    summary_headers = ["Theme", "Total Controls", "Fully Implemented", "Partially Implemented", "Not Implemented", "Compliance %"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws6.cell(row=9, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    themes = [("Organizational", 37), ("People", 8), ("Physical", 14), ("Technological", 34)]
    for row, (theme, count) in enumerate(themes, 10):
        ws6.cell(row=row, column=1, value=theme).border = THIN_BORDER
        ws6.cell(row=row, column=2, value=count).border = THIN_BORDER
        for col in range(3, 7):
            ws6.cell(row=row, column=col).border = THIN_BORDER
    
    ws6.cell(row=15, column=1, value="Gap Severity Distribution").font = Font(bold=True, size=14)
    severity_headers = ["Severity", "Count", "Percentage"]
    for col, header in enumerate(severity_headers, 1):
        cell = ws6.cell(row=16, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    severities = ["Critical", "High", "Medium", "Low"]
    for row, sev in enumerate(severities, 17):
        ws6.cell(row=row, column=1, value=sev)
    
    set_column_widths(ws6, [25, 18, 20, 22, 18, 15])
    
    # Save workbook
    wb.save("/Volumes/Data-Personal/Webpage/thehgtech/downloads/iso-27001-gap-analysis-template.xlsx")
    print("✅ Created: iso-27001-gap-analysis-template.xlsx")


def create_certification_toolkit():
    """Create ISO 27001:2022 Certification Toolkit"""
    wb = Workbook()
    
    # ========== Sheet 1: Project Plan ==========
    ws1 = wb.active
    ws1.title = "Project Plan"
    
    headers = ["Phase", "Task", "Description", "Owner", "Start Date", "End Date", 
               "Duration (Days)", "Status", "Dependencies", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws1, [12, 35, 45, 15, 12, 12, 15, 12, 20, 30])
    
    # Project phases and tasks
    tasks = [
        ("Phase 1", "Executive Sponsorship", "Secure CEO/CISO commitment and budget approval", "", "", "", "5", "", "-", ""),
        ("Phase 1", "Define ISMS Scope", "Document business units, locations, systems in scope", "", "", "", "10", "", "1", ""),
        ("Phase 1", "Gap Analysis", "Assess current state vs ISO 27001:2022 requirements", "", "", "", "20", "", "2", ""),
        ("Phase 1", "Project Planning", "Create detailed implementation timeline and RACI", "", "", "", "5", "", "3", ""),
        ("Phase 2", "Policy Development", "Create/update information security policies", "", "", "", "30", "", "4", ""),
        ("Phase 2", "Risk Assessment", "Identify assets, threats, vulnerabilities, calculate risks", "", "", "", "25", "", "4", ""),
        ("Phase 2", "Statement of Applicability", "Document control applicability and justifications", "", "", "", "10", "", "6", ""),
        ("Phase 2", "Control Implementation", "Implement controls based on risk treatment plan", "", "", "", "60", "", "7", ""),
        ("Phase 2", "Documentation", "Create procedures and work instructions", "", "", "", "40", "", "8", ""),
        ("Phase 2", "Training Program", "Develop and deliver security awareness training", "", "", "", "20", "", "5", ""),
        ("Phase 3", "Internal Audit Planning", "Define audit scope, select auditors, create schedule", "", "", "", "5", "", "9", ""),
        ("Phase 3", "Internal Audit Execution", "Conduct audits, document findings", "", "", "", "10", "", "11", ""),
        ("Phase 3", "Corrective Actions", "Address internal audit findings", "", "", "", "15", "", "12", ""),
        ("Phase 3", "Management Review", "Conduct formal management review meeting", "", "", "", "2", "", "13", ""),
        ("Phase 4", "CB Selection", "Select and contract certification body", "", "", "", "10", "", "14", ""),
        ("Phase 4", "Stage 1 Preparation", "Prepare documentation package for Stage 1", "", "", "", "10", "", "15", ""),
        ("Phase 4", "Stage 1 Audit", "Documentation review with certification body", "", "", "", "2", "", "16", ""),
        ("Phase 4", "Stage 1 Remediation", "Address Stage 1 findings", "", "", "", "20", "", "17", ""),
        ("Phase 5", "Stage 2 Preparation", "Final preparation for certification audit", "", "", "", "10", "", "18", ""),
        ("Phase 5", "Stage 2 Audit", "On-site certification audit", "", "", "", "5", "", "19", ""),
        ("Phase 5", "NC Remediation", "Address any non-conformities", "", "", "", "15", "", "20", ""),
        ("Phase 5", "Certification", "Receive ISO 27001 certificate", "", "", "", "5", "", "21", ""),
    ]
    
    for row, task in enumerate(tasks, 2):
        for col, value in enumerate(task, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
    
    from openpyxl.worksheet.datavalidation import DataValidation
    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,On Hold,Delayed"')
    ws1.add_data_validation(status_dv)
    status_dv.add("H2:H50")
    
    # ========== Sheet 2: Mandatory Documentation ==========
    ws2 = wb.create_sheet("Mandatory Documentation")
    
    headers2 = ["Clause", "Document Name", "Description", "Status", "Version", 
                "Approval Date", "Approver", "Location", "Review Date"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws2, [12, 35, 45, 12, 10, 12, 15, 30, 12])
    
    docs = [
        ("4.3", "ISMS Scope", "Document defining ISMS boundaries and applicability"),
        ("5.2", "Information Security Policy", "Top-level policy signed by management"),
        ("6.1.2", "Risk Assessment Methodology", "Process for identifying and assessing risks"),
        ("6.1.2", "Risk Assessment Results", "Output of risk assessment activities"),
        ("6.1.3", "Risk Treatment Plan", "Plan for addressing identified risks"),
        ("6.1.3", "Statement of Applicability", "Controls applicability with justifications"),
        ("6.2", "Information Security Objectives", "Measurable security objectives"),
        ("7.2", "Competence Evidence", "Training records, certifications, qualifications"),
        ("8.1", "Operational Planning Documents", "Procedures for operational processes"),
        ("9.2", "Internal Audit Program", "Audit schedule and procedures"),
        ("9.2", "Internal Audit Results", "Audit reports and findings"),
        ("9.3", "Management Review Records", "Meeting minutes and decisions"),
        ("10.1", "Nonconformity Records", "NC documentation and corrective actions"),
    ]
    
    for row, (clause, name, desc) in enumerate(docs, 2):
        ws2.cell(row=row, column=1, value=clause).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws2.cell(row=row, column=3, value=desc).border = THIN_BORDER
        for col in range(4, 10):
            ws2.cell(row=row, column=col).border = THIN_BORDER
    
    doc_status = DataValidation(type="list", formula1='"Not Started,Draft,In Review,Approved,Needs Update"')
    ws2.add_data_validation(doc_status)
    doc_status.add("D2:D50")
    
    # ========== Sheet 3: SoA Template ==========
    ws3 = wb.create_sheet("Statement of Applicability")
    
    headers3 = ["Control ID", "Control Name", "Applicable", "Justification", 
                "Implementation Status", "Implementation Method", "Reference Document"]
    
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws3, [12, 45, 12, 40, 20, 35, 30])
    
    # Add all 93 controls (reusing from gap analysis)
    controls = [
        ("A.5.1", "Policies for information security"),
        ("A.5.2", "Information security roles and responsibilities"),
        ("A.5.3", "Segregation of duties"),
        ("A.5.7", "Threat intelligence"),
        ("A.5.23", "Information security for use of cloud services"),
        ("A.6.3", "Information security awareness, education and training"),
        ("A.8.9", "Configuration management"),
        ("A.8.13", "Information backup"),
        ("A.8.16", "Monitoring activities"),
        ("A.8.28", "Secure coding"),
        # Add more...
    ]
    
    for row, (ctrl_id, ctrl_name) in enumerate(controls, 2):
        ws3.cell(row=row, column=1, value=ctrl_id).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=ctrl_name).border = THIN_BORDER
        for col in range(3, 8):
            ws3.cell(row=row, column=col).border = THIN_BORDER
    
    applicable_dv = DataValidation(type="list", formula1='"Yes,No"')
    impl_status_dv = DataValidation(type="list", formula1='"Implemented,Partially Implemented,Planned,Not Applicable"')
    ws3.add_data_validation(applicable_dv)
    ws3.add_data_validation(impl_status_dv)
    applicable_dv.add("C2:C200")
    impl_status_dv.add("E2:E200")
    
    # ========== Sheet 4: Internal Audit Checklist ==========
    ws4 = wb.create_sheet("Internal Audit Checklist")
    
    headers4 = ["Clause/Control", "Audit Question", "Evidence Required", 
                "Finding", "Evidence Reviewed", "Auditor Notes"]
    
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    set_column_widths(ws4, [15, 50, 40, 15, 35, 40])
    
    audit_questions = [
        ("4.1", "How does the organization determine external and internal issues relevant to the ISMS?", "Context analysis, SWOT, meeting minutes"),
        ("4.2", "How are interested parties and their requirements identified?", "Stakeholder register, requirements matrix"),
        ("5.1", "How does top management demonstrate commitment to the ISMS?", "Management review records, resource allocation"),
        ("5.2", "Is the information security policy documented, communicated, and available?", "Policy document, communication records"),
        ("6.1", "How are risks and opportunities identified and addressed?", "Risk register, risk assessment methodology"),
        ("7.2", "How is competence ensured for personnel affecting ISMS?", "Training records, job descriptions, certifications"),
        ("8.1", "How are operational processes planned and controlled?", "Procedures, work instructions, process documentation"),
        ("9.1", "How is ISMS performance evaluated?", "Metrics, KPIs, dashboards, reports"),
        ("9.2", "Is the internal audit program implemented?", "Audit schedule, audit reports, NC records"),
        ("9.3", "Is management review conducted at planned intervals?", "Meeting minutes, action items, decisions"),
        ("10.1", "How are nonconformities addressed?", "NC log, corrective action records, root cause analysis"),
    ]
    
    for row, (clause, question, evidence) in enumerate(audit_questions, 2):
        ws4.cell(row=row, column=1, value=clause).border = THIN_BORDER
        ws4.cell(row=row, column=2, value=question).border = THIN_BORDER
        ws4.cell(row=row, column=3, value=evidence).border = THIN_BORDER
        for col in range(4, 7):
            ws4.cell(row=row, column=col).border = THIN_BORDER
    
    finding_dv = DataValidation(type="list", formula1='"Conforming,Major NC,Minor NC,Observation,Not Audited"')
    ws4.add_data_validation(finding_dv)
    finding_dv.add("D2:D100")
    
    # ========== Sheet 5: Management Review Template ==========
    ws5 = wb.create_sheet("Management Review")
    
    ws5.cell(row=1, column=1, value="ISO 27001 Management Review").font = Font(bold=True, size=16)
    
    ws5.cell(row=3, column=1, value="Meeting Details").font = Font(bold=True, size=12)
    details = ["Date:", "Attendees:", "Chair:", "Minutes By:"]
    for row, detail in enumerate(details, 4):
        ws5.cell(row=row, column=1, value=detail).font = Font(bold=True)
    
    ws5.cell(row=9, column=1, value="Required Inputs (ISO 27001 Clause 9.3)").font = Font(bold=True, size=12)
    
    inputs = [
        "a) Status of actions from previous management reviews",
        "b) Changes in external and internal issues relevant to the ISMS",
        "c) Feedback on information security performance, including: nonconformities and corrective actions, monitoring and measurement results, audit results, fulfillment of objectives",
        "d) Feedback from interested parties",
        "e) Results of risk assessment and status of risk treatment plan",
        "f) Opportunities for continual improvement",
    ]
    
    for row, inp in enumerate(inputs, 10):
        ws5.cell(row=row, column=1, value=inp)
        ws5.cell(row=row, column=2, value="Status/Notes:")
    
    ws5.cell(row=17, column=1, value="Required Outputs").font = Font(bold=True, size=12)
    outputs = [
        "Decisions related to continual improvement opportunities",
        "Decisions related to changes to the ISMS",
        "Resource needs",
    ]
    
    for row, out in enumerate(outputs, 18):
        ws5.cell(row=row, column=1, value=out)
        ws5.cell(row=row, column=2, value="Decision/Action:")
    
    ws5.cell(row=22, column=1, value="Action Items").font = Font(bold=True, size=12)
    action_headers = ["Action", "Owner", "Due Date", "Status"]
    for col, header in enumerate(action_headers, 1):
        cell = ws5.cell(row=23, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    set_column_widths(ws5, [60, 30, 15, 15])
    
    # ========== Sheet 6: Audit Preparation ==========
    ws6 = wb.create_sheet("Audit Preparation")
    
    ws6.cell(row=1, column=1, value="Stage 1 & Stage 2 Audit Preparation Checklist").font = Font(bold=True, size=14)
    
    headers6 = ["Item", "Description", "Ready", "Location/Notes"]
    for col, header in enumerate(headers6, 1):
        cell = ws6.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    stage1_items = [
        ("ISMS Scope", "Documented and approved ISMS scope statement"),
        ("Information Security Policy", "Signed, dated, communicated to all staff"),
        ("Risk Assessment", "Methodology documented, assessment completed"),
        ("Risk Treatment Plan", "Plan documented with owners and timelines"),
        ("Statement of Applicability", "All 93 controls addressed with justifications"),
        ("Internal Audit Results", "Audit completed, findings documented"),
        ("Management Review", "Review conducted, minutes available"),
        ("Objectives and Metrics", "Security objectives defined and measured"),
        ("Competence Records", "Training records for key personnel"),
        ("Organization Chart", "Showing ISMS roles and responsibilities"),
    ]
    
    ws6.cell(row=2, column=1, value="Stage 1 Requirements").font = Font(bold=True)
    for row, (item, desc) in enumerate(stage1_items, 4):
        ws6.cell(row=row, column=1, value=item).border = THIN_BORDER
        ws6.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws6.cell(row=row, column=3).border = THIN_BORDER
        ws6.cell(row=row, column=4).border = THIN_BORDER
    
    ws6.cell(row=15, column=1, value="Stage 2 Requirements").font = Font(bold=True)
    
    stage2_items = [
        ("All Stage 1 items", "All items from Stage 1 still current"),
        ("Stage 1 findings addressed", "Evidence of closure for any Stage 1 findings"),
        ("Operating records", "3+ months of ISMS operational evidence"),
        ("Incident records", "Incident log with response evidence"),
        ("Access review records", "Evidence of periodic access reviews"),
        ("Backup/restore tests", "Recent backup restore test results"),
        ("Vulnerability assessments", "Recent vulnerability scan results"),
        ("Training records", "Current training completion for all staff"),
        ("Audit rooms", "Quiet room for interviews prepared"),
        ("Key personnel availability", "Ensure all process owners available"),
    ]
    
    for row, (item, desc) in enumerate(stage2_items, 16):
        ws6.cell(row=row, column=1, value=item).border = THIN_BORDER
        ws6.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws6.cell(row=row, column=3).border = THIN_BORDER
        ws6.cell(row=row, column=4).border = THIN_BORDER
    
    set_column_widths(ws6, [30, 50, 10, 35])
    
    ready_dv = DataValidation(type="list", formula1='"Yes,No,Partial"')
    ws6.add_data_validation(ready_dv)
    ready_dv.add("C4:C30")
    
    # Save workbook
    wb.save("/Volumes/Data-Personal/Webpage/thehgtech/downloads/iso-27001-certification-toolkit.xlsx")
    print("✅ Created: iso-27001-certification-toolkit.xlsx")


if __name__ == "__main__":
    print("Creating ISO 27001:2022 Templates for TheHGTech...")
    create_gap_analysis_template()
    create_certification_toolkit()
    print("\n🎉 All templates created successfully in /downloads/")
